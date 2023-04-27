import json
import jsonpath
import requests
import openpyxl

class Common:

    def __init__(self, FileNamePath, SheetName):
        global ex, sh
        ex = openpyxl.load_workbook(FileNamePath)
        sh = ex[SheetName]

    def fetch_row_count(self):
        rows = sh.max_row
        return rows

    def fetch_column_count(self):
        col = sh.max_column
        return col

    def fetch_key_names(self):
        c = sh.max_column #buscar cuantas hay en la hoja, se retornan 4 porque el libro tiene 4 columnas
        li=[] # se crea una lista vacia
        for i in range(1,c+1): # se ejecuta el ciclo de 1-5
            cell = sh.cell(row=1, column=i) # se recorre desde la primera fila y la columna es el valor i del ciclo
            li.insert(i-1, cell.value) # se inserta en la lista el valor
        return li

    def update_request_with_data(self, rowNumber, jsonRequest, keyList):
        c = sh.max_column
        for i in range(1,c+1): # se recorren todas las columnas xq cada fila tiene multiples columnas
           cell = sh.cell(rowNumber, column=i) #Buscar data de una columna en particular
           jsonRequest[keyList[i-1]] = cell.value
        return jsonRequest




